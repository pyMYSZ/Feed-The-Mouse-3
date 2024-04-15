"""
Feed The Mouse 3 
Author:    MYSZ - https://github.com/pyMYSZ
type:      game - pygame
version:   3.0.0
date:      25-01-2024
"""

import random
import math
import pygame
import sys
import config.settings as gs
import pandas as pd
import openpyxl
import os
import subprocess
import time
import json
import webbrowser

pygame.init()

# Debugs
debug_mouse_pos = False  # True = print mouse pos. every 1 sek.
debug_print_image_info = False  # True = print all loaded images
debug_collision = False  # True = draw all object rect
mouse_counter = 0  # to update mouse pos.

# region 1. Start value
intro_screen = True  # Feed The Mouse 3   image
loading_screen = False  # Loading...         image
menu_screen = True  # True = player in menu / False = player in game
game_running = True  # False = quit game

# Menu
menu_start = True
menu_settings = False
menu_score = False
menu_licenses = False
menu_controls = False
menu_shop = False
menu_info = False
editing_key = None
load_high_scores = False

COLORS = gs.create_colors_list(size=30)
color_index = 0
menu_color = gs.YELLOW
select_color = gs.WHITE

start_text = ""
high_score_text = ""
settings_text = ""
shop_text = ""
controls_menu_typ = 1  # 0: save,  1: load,  2: clear
volume = 0
volume_temp = volume
volume_changed = False
volume_temp_color = gs.BLACK
cash_ratio = 0

# Game
skin_not_owned = False  # Start with basic skin if player use not owned skin
save_score = True  # FLAG: Save score to Excel file only one times
save_score2 = True  # FLAG: Save score to Excel file only one times
game_screen = False  # True = game window  / False = menu window
game_over = False
game_paused = False
music_playing = True

info_image_index = 0
pause_game_image_index = 0  # random.randint(0, 2)
high_score_image_index = 1  # 1 = TOP 10, 2 = TOP 30
high_score_color_index = 0
high_score_color_list = [gs.BLACK, gs.WHITE, gs.YELLOW]
number_of_players = 1

last_update_time = time.time()  # to update cheese speed
cheese_speed = 0.00  # current cheese speed
show_cheese_speed: bool = True  # True = show cheese speed in game HUD
# endregion


# region 2. Create Game Window
logo_image = pygame.image.load('./config/images/logo.png')
image_intro = pygame.image.load('./config/images/menu/MYSZ_games_intro.png')
intro_sounds = pygame.mixer.Sound('./config/music/intro.mp3')
sound_FeedTheMouse = pygame.mixer.Sound('./config/music/FeedTheMouse3.mp3')
pygame.display.set_caption("Feed the Mouse 3 ")
pygame.display.set_icon(logo_image)

# Play Intro music
intro_sounds.play()
sound_FeedTheMouse.play()

# Create window game
WG = pygame.display.set_mode((gs.WINDOW_WIDTH, gs.WINDOW_HEIGHT))
if intro_screen:
    WG.fill(gs.BLACK)
    WG.blit(image_intro, (225, 25))
    pygame.display.flip()

# Set FPS and clock
FPS = gs.FPS
clock = pygame.time.Clock()

# Cursor
current_cursor = pygame.image.load('./config/images/cursor.png').convert_alpha()
pygame.mouse.set_cursor((8, 8), (0, 0), (0, 0, 0, 0, 0, 0, 0, 0), (0, 0, 0, 0, 0, 0, 0, 0))


# endregion


# region 3. Load Data
# Save data to JSON file
def save_data_to_file(data, slot_number, filename_prefix='./config/data/controls_slot_'):
    """ data - file to save """
    filename = f'{filename_prefix}{slot_number}.json'
    with open(filename, 'w') as file:
        json.dump(data, file, indent='\n')
        print(f"Save data to {filename_prefix}{slot_number}.")


# Load data from JSON file
def load_data_from_file(slot_number, filename_prefix='./config/data/controls_slot_'):
    filename = f'{filename_prefix}{slot_number}.json'
    try:
        with open(filename, 'r') as file:
            print(f"load data from  {filename_prefix}{slot_number}.")
            return json.load(file)
    except FileNotFoundError:
        print(f"Data {filename_prefix}{slot_number} not exist")
        return None


# Clear data in JSON file
def restart_data_file(slot_number, filename_prefix='./config/data/controls_slot_'):
    """ restores initial settings"""
    filename = f'{filename_prefix}{slot_number}.json'
    try:
        os.remove(filename)
        print(f"Clear data in  {slot_number}.")
    except FileNotFoundError:
        print(f"Data {slot_number} not exist")
    temp_controls = load_data_from_file(slot_number=0)
    save_data_to_file(data=temp_controls, slot_number=slot_number)
    default_settings_data = load_data_from_file(slot_number=slot_number)

    return default_settings_data


# Restore TopList
def restore_top_list(folder_path='./config/data/', output_path='./config/data/score.xlsx'):
    """ restore top list """
    # Load all files
    files = [file for file in os.listdir(folder_path) if file.startswith('score_backUp_') and file.endswith('.xlsx')]

    # Add all data
    merged_data = pd.DataFrame()
    for file in files:
        file_path = os.path.join(folder_path, file)
        df = pd.read_excel(file_path)
        merged_data = pd.concat([merged_data, df], ignore_index=True)

    try:
        existing_data = pd.read_excel(output_path)
        merged_data = pd.concat([existing_data, merged_data], ignore_index=True)
    except FileNotFoundError:
        print('Cant restore topList')

    # Save all data in one file
    merged_data.to_excel(output_path, index=False)

    # Delete old data
    for file in files:
        file_path = os.path.join(folder_path, file)
        os.remove(file_path)


# Load controls and shop data
controls = load_data_from_file(slot_number=0, filename_prefix='./config/data/controls_slot_')
shop_data = load_data_from_file(slot_number=1, filename_prefix='./config/data/costume_shop_')
money = int(shop_data["money"])
costumes = shop_data["costumes"]

# Load game settings data
game_settings = load_data_from_file(slot_number=0, filename_prefix='./config/data/settings_')
player1_skin_image_index = game_settings["player_1"]
player2_skin_image_index = game_settings["player_2"]
volume = game_settings["volume"]

# Load Score Table
license_file_path = './config/data/Licenses.xlsx'
excel_file_path = './config/data/score.xlsx'
if os.path.exists(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Score", "Date", "Cheese"])
    wb.save(excel_file_path)

# Load High Score List Data
wb = openpyxl.load_workbook(excel_file_path)
ws = wb.active
sorted_results = sorted(ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row),
                        key=lambda var_x: var_x[0].value, reverse=True)
top10_score_list = sorted_results[:10]  # top 10 best score
top30_score_list = sorted_results[:30]  # top 30 best score

# Load Sources List Data
wbs = openpyxl.load_workbook(license_file_path)
wss = wbs.active

sources_name = []
sources_link = []
sources_typ = []
for row in wss.iter_rows(min_row=2, max_row=wss.max_row, min_col=3, max_col=5):
    sources_name.append(row[0].value)
    sources_link.append(row[1].value)
    sources_typ.append(row[2].value)

licenses_index = 0
licenses_max_index = len(sources_typ) - 1


# endregion


# region 4.0. Loading images

class SpriteImage:
    def __init__(self, image):
        """ image: pygame.image.load('image_path') """
        self.sheet = image
        self.width = self.sheet.get_width()
        self.height = self.sheet.get_height()

    def get_image_one_sheet(self, amount, img_idx=0, scale=1, x=0, y=0):
        """ Return single image (index) from one-row sprites sheet """
        width = self.width // amount
        height = self.height
        image = pygame.Surface((width, height)).convert_alpha()
        image.blit(self.sheet, (x, y), (img_idx * width, 0, width, height))
        image = pygame.transform.scale(image, (width * scale, height * scale))
        image.set_colorkey('black')  # make image transparent

        return image

    def get_image_x_sheet(self, columns, rows, column_index=0, row_index=0, scale=1, angle=0, x=0, y=0, color='black'):
        """ Return single image (row_idx, column_index) from sprites sheet """
        width = self.width // columns
        height = self.height // rows
        image = pygame.Surface((width, height)).convert_alpha()
        image.blit(self.sheet, (x, y), (column_index * width, row_index * height, width, height))
        image = pygame.transform.scale(image, (width * scale, height * scale))
        image = pygame.transform.rotate(image, angle)
        image.set_colorkey(color)
        return image

    def get_images_list(self, amount_x, amount_y, scale=1, angle=0, flip_x=False, flip_y=False, x=0, y=0,
                        color='black'):
        """ Return images list from sprite graphic """
        width = self.width // amount_x
        height = self.height // amount_y
        image_list = []

        for row in range(amount_y):
            for col in range(amount_x):
                image = pygame.Surface((width, height)).convert_alpha()
                image.blit(self.sheet, (x, y), (col * width, row * height, width, height))
                image = pygame.transform.scale(image, (width * scale, height * scale))
                image = pygame.transform.rotate(image, angle)
                image = pygame.transform.flip(image, flip_x, flip_y)
                image.set_colorkey(color)
                image_list.append(image)
        return image_list


image_buttons_score_menu = pygame.image.load('./config/images/menu/buttons_score.png')
image_buttons_start_menu = pygame.image.load('./config/images/menu/buttons_start.png')
image_buttons_info_menu = pygame.image.load('./config/images/menu/buttons_info.png')
image_player_locked = pygame.image.load('./config/images/menu/lock.png')
image_info = pygame.image.load('./config/images/menu/info.png')
image_coca_cola = pygame.image.load('./config/images/game/cola.png')

sheet_image_egg = SpriteImage(pygame.image.load(f'./config/images/sheet_img/eggs_crush.png'))
img_list_egg = sheet_image_egg.get_images_list(7, 1, 0.85)
image_egg = gs.scale_image(pygame.image.load('./config/images/game/egg.png'), 0.85)

sheet_image_water = SpriteImage(pygame.image.load(f'./config/images/sheet_img/water_crush.png'))
img_list_water = sheet_image_water.get_images_list(4, 2, 0.65)
image_water = gs.scale_image(pygame.image.load('./config/images/game/water.png'), 0.65)

sheet_image_poison = SpriteImage(pygame.image.load(f'./config/images/sheet_img/poison_crush.png'))
img_list_poison = sheet_image_poison.get_images_list(4, 2, 0.65)
image_poison = gs.scale_image(pygame.image.load('./config/images/game/poison.png'), 0.65)

sheet_image_trap = SpriteImage(pygame.image.load(f'./config/images/sheet_img/mousetrap.png'))
img_list_mousetrap = sheet_image_trap.get_images_list(3, 4, 0.45)

sheet_image_redbull = SpriteImage(pygame.image.load(f'./config/images/sheet_img/redbull.png'))
img_list_soda = sheet_image_redbull.get_images_list(20, 1, 1)

sheet_image_medicineKit = SpriteImage(pygame.image.load(f'./config/images/sheet_img/medicine_kit.png'))
img_list_medicineKit = sheet_image_medicineKit.get_images_list(3, 4, 0.45)

# Backgrounds images
image_loading = pygame.image.load('./config/images/menu/MYSZ_games_loading.png')
image_menu_start = pygame.image.load('./config/images/menu/MYSZ_games_start.png')
image_menu_settings = pygame.image.load('./config/images/menu/MYSZ_games_settings.png')
image_menu_licenses = pygame.image.load('./config/images/menu/MYSZ_games_licenses.png')
image_menu_controls = pygame.image.load('./config/images/menu/MYSZ_games_controls.png')
image_menu_players = pygame.image.load('./config/images/menu/MYSZ_games_players.png')
image_menu_shop = pygame.image.load('./config/images/menu/MYSZ_games_shop.png')

rect_top_right = (gs.WINDOW_WIDTH - 325, 0)
rect_top_left = image_menu_settings.get_rect(topleft=(0, 0))
rect_center = image_loading.get_rect(topleft=((gs.WINDOW_WIDTH - 750) // 2, 0))

# GameOver
image_game_over_one_player = pygame.image.load('./config/images/game/game_over_hud.png')
image_game_over_two_players = pygame.image.load('./config/images/game/game_over_hud2.png')

# Create images list from file ./classes
IMAGES = gs.ImageList("./config/images/classes")
img_list_pause_game = IMAGES.get_list('pauseMouse', '_', debug_print_image_info)
img_list_player = IMAGES.get_list('mouse', '_')
img_list_cheese = IMAGES.get_list('cheese', '_')
img_list_game_over = IMAGES.get_list('gameOver', '_')
img_list_high_score = IMAGES.get_list('highScore', '_')
img_list_how_to_play = IMAGES.get_list('HowToPlay', '_')
# endregion

# region 4.1. Loading sounds
# Background music
pygame.mixer.init()
pygame.mixer.music.load('./config/music/bg.wav')
pygame.mixer_music.play(-1, 0.0, 12500)
pygame.mixer.music.set_volume(volume / 100)


def music_player():
    """ Pause or unpause background music """
    global music_playing
    music_playing = not music_playing
    return pygame.mixer.music.unpause() if music_playing else pygame.mixer.music.pause()


# Sounds effect
sound_gameOver = pygame.mixer.Sound('./config/music/game_over.wav')
sound_buy_skin = pygame.mixer.Sound('./config/music/shop.mp3')
sound_trap_start = pygame.mixer.Sound('./config/music/trap.mp3')
sound_trap_hit = pygame.mixer.Sound('./config/music/trap_hit.mp3')
sound_medicine_hit = pygame.mixer.Sound('./config/music/medicine.mp3')
click_sound = pygame.mixer.Sound('./config/music/click.mp3')


# endregion


# region BONUS - CLASSES IMPORT - one file
class ButtonText:
    def __init__(self, x, y, width, height, text, font, default_color, hover_color, frame=True, frame_width=2):
        self.rect = pygame.Rect(x, y, width, height)
        self.text = text
        self.font = font
        self.default_color = default_color
        self.hover_color = hover_color
        self.frame_width = frame_width
        self.framed = frame
        self.current_color = default_color
        self.clicked = False
        self.counter = 0

    def draw(self, surface):
        action = False
        pos = pygame.mouse.get_pos()

        if self.rect.collidepoint(pos):
            self.current_color = self.hover_color if not self.clicked else (220, 10, 10)
            if pygame.mouse.get_pressed()[0] and not self.clicked:
                action = True
                click_sound.play()
                self.clicked = True
        else:
            self.current_color = self.default_color if not self.clicked else (220, 10, 10)

        # Draw frame
        if self.framed:
            # pygame.draw.rect(surface, self.current_color, self.rect, self.frame_width)
            pygame.draw.rect(surface, (255, 255, 255), self.rect, self.frame_width)

        # Draw Text in frame
        text_surface = self.font.render(self.text, True, self.current_color)
        text_rect = text_surface.get_rect(center=self.rect.center)
        surface.blit(text_surface, text_rect)

        # Restart button
        self.counter += 1
        self.restart() if self.counter >= 200 else None

        return action

    def restart(self):
        self.counter = 0
        self.clicked = False

    def update(self, color):
        self.default_color = color


class ButtonRect:
    def __init__(self, top_left, bottom_right, default_color, hover_color, time=200, action=False, link=False,
                 index=None, description=""):
        self.rect = pygame.Rect(top_left[0], top_left[1], bottom_right[0] - top_left[0], bottom_right[1] - top_left[1])
        self.default_color = default_color
        self.hover_color = hover_color
        self.current_color = default_color
        self.clicked = False
        self.counter = 0
        self.time = time
        self.make_unfreeze = action
        self.freeze = False
        self.link = link
        self.index = index
        self.description = description

    def draw(self, surface):
        action = False

        if self.mouse_in():
            self.current_color = self.hover_color
            if pygame.mouse.get_pressed()[0] and not self.clicked:
                action = True
                click_sound.play()
                self.clicked = True
                self.freeze = True if self.make_unfreeze else False
        else:
            self.current_color = self.default_color if not self.clicked else (220, 10, 10)

        # Draw Rect
        pygame.draw.rect(surface, self.current_color, self.rect)

        # Restart button
        self.counter += 1
        self.restart() if self.counter >= self.time else None

        return action

    def restart(self):
        self.counter = 0
        self.clicked = False
        self.freeze = False

    def update(self, color):
        self.default_color = color

    def mouse_in(self):
        action = False
        pos = pygame.mouse.get_pos()

        if self.rect.collidepoint(pos):
            action = True

        return action


class GameState:
    def __init__(self):
        """ class stores global game values for all other classes """
        self.TOTAL_CHEESE = 0            # total amount of cheese (normal & bonus) in one game
        self.TOTAL_FAKE = 0              # total amount of FAKE cheese in one game
        self.MISS_IN_ROW = 0             # number of cheeses missed in a row
        self.FAKE_IN_ROW = 0             # number of FAKE cheeses caught in a row
        self.CATCH_IN_ROW = 0            # number of cheeses caught in a row
        self.SPAWN_POS_LIST = [(0, 0)]   # list of the last items(cheese) pos. - cheeses can't spawn next to each other
        self.TRAP_POS_LIST = [(0, 0)]    # list of the last mousetrap pos. -  can't spawn next to each other
        self.LIST_SIZE = 3               # maximum size of list of recent respawn

    def restart_cheese(self):
        """ restart of cheese stats (new game) """
        self.TOTAL_CHEESE = 0
        self.TOTAL_FAKE = 0
        self.MISS_IN_ROW = 0
        self.FAKE_IN_ROW = 0
        self.SPAWN_POS_LIST = []


class Player:
    def __init__(self, game_state: GameState, x: int, y: int, images: list, typ: int = 1):
        self.game_state = game_state
        self.typ: int = typ                   # 1: Player1, 2: Player2
        self.game_over: int = 3               # Game Over Typ: 0 - hp, 1 - time, 2 - poison,  3 - new game
        self.dead: bool = False               # Check is player dead

        # player achievements
        self.score = 0                        # total score
        self.cheese_amount = 0                # total amount of standard cheese caught
        self.cheese_bonus_amount = 0          # total amount of bonus cheese caught
        self.cheese_fake_amount = 0           # total amount of fake cheese caught

        # player movement
        self.velocity = gs.PLAYER_VELOCITY
        self.start_velocity = gs.PLAYER_VELOCITY
        self.velocity_normal = gs.PLAYER_VELOCITY
        self.velocity_boost = gs.PLAYER_VELOCITY * gs.PLAYER_BOOST_SCALE
        self.horizontal = False     # player move x direction
        self.vertical = False       # player move y direction
        self.slowed = False         # player is slowed
        self.boosted = False        # player using boost
        self.penetration = True     # player can penetrate vertical walls
        self.miss_3_in_row = True
        self.miss_5_in_row = True
        self.miss_10_in_row = True

        # player statistics
        self.health = gs.PLAYER_MAX_HP
        self.max_hp = gs.PLAYER_MAX_HP
        self.boost_lvl = gs.PLAYER_MAX_BOOST
        self.max_boost = gs.PLAYER_MAX_BOOST
        self.wraps = gs.PLAYER_MAX_WRAPS
        self.max_wraps = gs.PLAYER_MAX_WRAPS
        self.poison = 0
        self.max_poison = gs.PLAYER_MAX_POISON

        # time settings
        self.slow_time = pygame.time.get_ticks()
        self.game_time = 0                 # sek.
        self.time_left = gs.ROUND_TIME     # sek.
        self.frame_count = 0               # to update time

        # sounds settings (cheese eating)
        self.sound_01 = pygame.mixer.Sound('./config/music/cheese.wav')            # normal cheese
        self.sound_02 = pygame.mixer.Sound('./config/music/cheese_bonus.mp3')      # bonus cheese
        self.sound_03 = pygame.mixer.Sound('./config/music/cheese_fake.mp3')       # moldy cheese
        self.sound_04 = pygame.mixer.Sound('./config/music/cheese_fake_big.mp3')   # big moldy cheese

        # image settings
        self.image_list = images
        self.max_index = len(images)
        self.image_index = 0
        self.image = self.image_list[self.image_index]
        self.rect_image = self.image.get_rect(center=(x, y))

        # collision settings
        rect_dx = 10
        rect_dy = 10
        new_width = self.image.get_width() - 2 * rect_dx
        new_height = self.image.get_height() - 2 * rect_dy
        self.rect = pygame.Rect(x + rect_dx, y + rect_dy, new_width, new_height)
        self.rect.center = self.rect_image.center

    def update(self, cheese_group):
        """ checks collisions with cheese and edges, values update & control: speed, time, ... """
        # update time var.
        self.frame_count += 1
        if self.frame_count == gs.FPS:
            self.time_left -= 1
            self.frame_count = 0
            if not self.dead:
                self.game_time += 1

        # update image with collision rect
        self.rect_image.centerx = self.rect.centerx
        self.rect_image.centery = self.rect.centery - 5

        # check that the value has not exceeded the maximum
        self.health = min(self.max_hp, self.health)
        self.boost_lvl = max(-5, min(self.max_boost, self.boost_lvl))
        self.wraps = min(self.max_wraps, self.wraps)
        self.score = max(0, self.score)

        # check for Game Over
        if self.health <= 0:
            self.dead = True
            self.game_over = 0
        elif self.time_left <= 0:
            self.dead = True
            self.game_over = 1
        elif self.poison >= self.max_poison:
            self.dead = True
            self.game_over = 2

        # check collision with cheese & unlock cheese
        for cheese in cheese_group:
            if self.rect.colliderect(cheese.rect) and not self.dead:
                if cheese.fake:
                    self.score += math.ceil(cheese.score // 2.5 + 1)
                    self.poison += math.ceil(cheese.score // 4.5 + 1)
                    self.cheese_fake_amount += 1
                    self.game_state.FAKE_IN_ROW += 1
                    self.health -= self.game_state.FAKE_IN_ROW + self.game_state.TOTAL_FAKE // 10
                    self.sound_03.play() if cheese.typ < 5 else self.sound_04.play()
                    self.game_state.CATCH_IN_ROW = -1
                    cheese.restart()
                else:
                    self.game_state.MISS_IN_ROW = 0
                    self.score += cheese.score
                    if cheese.typ == 2:
                        self.cheese_bonus_amount += 1
                        self.time_left += 1.50
                        self.sound_02.play()
                        self.update_speed(0.025)
                    else:
                        self.cheese_amount += 1
                        self.time_left += 0.25
                        self.update_speed(0.005)
                        self.sound_01.play()
                    self.game_state.CATCH_IN_ROW += 1

                    # add extra time
                    if self.game_state.CATCH_IN_ROW >= 5:
                        self.game_state.CATCH_IN_ROW = 0
                        self.time_left += 5

                    # restart
                    self.miss_3_in_row = True
                    self.miss_5_in_row = True
                    self.miss_10_in_row = True
                    cheese.restart()

            # unlock cheese
            if cheese.locked and self.score > cheese.unlock_score:
                cheese.locked = False

        # update boost
        if self.boost_lvl > 5:
            self.boost_lvl += 1 if (self.horizontal or self.vertical) else 2
        else:
            self.boost_lvl += 0.04

        # remove slow
        if self.slowed and pygame.time.get_ticks() - self.slow_time > gs.PLAYER_SLOW_TIME:
            self.slowed = False

        # lost hp after multiple cheese miss
        if self.game_state.MISS_IN_ROW >= 10:
            if self.miss_10_in_row:
                self.miss_10_in_row = False
                self.health -= 4
            else:
                self.health -= 0.06

        if self.game_state.MISS_IN_ROW >= 5:
            if self.miss_5_in_row:
                self.miss_5_in_row = False
                self.health -= 4
            else:
                self.health -= 0.02

        if self.game_state.MISS_IN_ROW >= 3 and self.miss_3_in_row:
            self.miss_3_in_row = False
            self.health -= 2

    def draw(self, surface):
        """ draw player images if not dead"""
        if not self.dead:
            surface.blit(self.image, self.rect_image)
            if self.penetration:
                surface.blit(self.image, self.rect_image.move(gs.WINDOW_WIDTH, 0))
                surface.blit(self.image, self.rect_image.move(-gs.WINDOW_WIDTH, 0))

    def move(self, controls: dict):
        """ update player position after move (key press) """
        self.horizontal = False
        self.vertical = False
        keys = pygame.key.get_pressed()

        if not self.dead:
            if keys[controls[f'LEFT_{self.typ}']]:
                if (not self.penetration and self.rect.left > 0) or self.penetration:
                    self.horizontal = True
                    self.rect.x -= self.velocity / math.sqrt(self.horizontal + self.vertical)
            if keys[controls[f'RIGHT_{self.typ}']]:
                if (not self.penetration and self.rect.right < gs.WINDOW_WIDTH) or self.penetration:
                    self.horizontal = True
                    self.rect.x += self.velocity / math.sqrt(self.horizontal + self.vertical)
            if keys[controls[f'UP_{self.typ}']] and self.rect.top > 0:
                self.vertical = True
                self.rect.y -= self.velocity / math.sqrt(self.horizontal + self.vertical)
            if keys[controls[f'DOWN_{self.typ}']] and self.rect.bottom < gs.HUD2:
                self.vertical = True
                self.rect.y += self.velocity / math.sqrt(self.horizontal + self.vertical)

            # Penetration through vertical walls
            if self.penetration and self.wraps >= 1:
                if self.rect.left < -gs.PLAYER_SIZE:
                    self.rect.right = gs.WINDOW_WIDTH
                    self.wraps -= 1
                elif self.rect.right > gs.WINDOW_WIDTH + gs.PLAYER_SIZE:
                    self.rect.left = 0
                    self.wraps -= 1
                if self.wraps < 1:
                    self.penetration = False

            # Engage Boost
            if keys[controls[f'BOOST_{self.typ}']] and self.boost_lvl > 5 and not self.slowed:
                self.velocity = self.velocity_boost if not self.slowed else 4.25
                self.boost_lvl -= 4
            else:
                self.velocity = self.velocity_normal if not self.slowed else 2.85

    def restart(self, img_index: int = None):
        """ restart Player settings to defaults values"""
        self.dead = False
        self.penetration = True
        self.slowed = False
        self.game_over = 3
        self.slow_time = pygame.time.get_ticks()
        self.time_left = gs.ROUND_TIME
        self.game_time = 0
        self.frame_count = 0
        self.score = 0
        self.cheese_amount = 0
        self.cheese_bonus_amount = 0
        self.cheese_fake_amount = 0
        self.poison = 0
        self.health = self.max_hp
        self.boost_lvl = self.max_boost
        self.wraps = self.max_wraps
        self.velocity_normal = self.start_velocity
        self.rect.center = (400 * self.typ, gs.WINDOW_HEIGHT - 150)
        self.miss_3_in_row = False
        self.miss_5_in_row = False
        self.miss_10_in_row = False
        if img_index is not None:
            self.image_index = img_index

    def draw_hud(self, surface, offset_x=0, bar_width=240, one_player=True):
        """ draws the players statistics with the graphics in game window """
        # Draw HP & boost BAR
        hp_ratio = self.health / self.max_hp
        boost_ratio = (self.max_boost-self.boost_lvl) / self.max_boost

        pygame.draw.rect(surface, gs.RED, (offset_x + 20, 650, bar_width * hp_ratio, 50))
        pygame.draw.rect(surface, gs.BLUE, (offset_x + 15, 705, bar_width + 10, 40))
        pygame.draw.rect(surface, gs.BLACK, (offset_x + 20, 705, bar_width * boost_ratio + 5, 40))

        gs.draw_text(surface=surface, text=f'{int(self.health / self.max_hp * 100)}%', x=140 + offset_x, y=675,
                     font=gs.FONT_32, center=True)
        gs.draw_text(surface=surface, text=f'{max(0, int(min(self.boost_lvl, self.max_boost)))}/{self.max_boost}',
                     x=140 + offset_x, y=727, font=gs.FONT_21, center=True)

        # Draw images
        img_name = 'player_bar.png' if one_player else f'bar_{self.typ}.png'
        gs.draw_image_by_path(surface, f'./config/images/game/{img_name}', 0, gs.HUD2, 1)

        # Draw Player Stats (single-player)
        if one_player:
            gs.draw_text(surface, f'{self.velocity:.2f}', 315, 730, gs.FONT_25, center=True)
            gs.draw_text(surface, f'{int(self.wraps)}/{self.max_wraps}', 385, 730, gs.FONT_25, center=True)
            gs.draw_text(surface, f'{(self.cheese_amount + self.cheese_bonus_amount)}', 465, 730,
                         gs.FONT_25, center=True)
            gs.draw_text(surface, f'{self.score}', 550, 730, gs.FONT_25, center=True)
            gs.draw_text(surface, f'{int(self.time_left) // 60:02d}:{int(self.time_left)  % 60:02d}', 745,
                         730, gs.FONT_25, center=True)

        # Draw Player Stats (multi-player)
        else:
            # draw time left, only once
            if self.typ == 1:
                gs.draw_text(surface, f'{int(self.time_left) // 60:02d}:{int(self.time_left) % 60:02d}', 600,
                             730, gs.FONT_25, center=True)

            # stats text value
            stats_x = 400 + 395 * (self.typ - 1)
            if self.dead:
                text_speed = ' - '
                text_cheese = ' - '
                text_wraps = ' - '
            else:
                text_speed = f'{self.velocity:.2f}'
                text_cheese = f'{(self.cheese_amount + self.cheese_bonus_amount)}'
                text_wraps = f'{int(self.wraps)}/{self.max_wraps}'

            # draw stats
            gs.draw_text(surface, text_speed, stats_x, 665, gs.FONT_25, center=True)
            gs.draw_text(surface, text_cheese, stats_x, 695, gs.FONT_25, center=True)
            gs.draw_text(surface, text_wraps, stats_x, 725, gs.FONT_25, center=True)

        # Draw HUD Line
        pygame.draw.line(surface, gs.WHITE, (0, gs.HUD2), (gs.WINDOW_WIDTH, gs.HUD2), 2)

    def update_image(self, img_index):
        """ changes player's image after selecting a skin  """
        self.image_index = img_index
        self.image = self.image_list[self.image_index]

    def update_time(self, player2):
        """ in multiplayer TIME LEFT is displayed jointly for both players """
        self.time_left = max(self.time_left, player2.time_left)

    def update_speed(self, speed):
        """ updates the player's speed when bonuses are added """
        self.velocity_normal += speed
        self.velocity_boost = self.velocity_normal * gs.PLAYER_BOOST_SCALE


class Cheese(pygame.sprite.Sprite):
    def __init__(self, game_state: GameState, x: int, y: int, images: list, typ: int = 0,
                 bonus_chance: int = 0, fake_chance: int = 0, unlock_score: int = 0):
        """
        - typ: 0,1=normal / 2=big(bonus) / 3,4=fake / 5=big & fake
        - extra_distance: additional distance (time) between restarts (reappearances) = random(min, max) from settings
        - bonus_chance / fake chance: Percentage chance to drop bonus/ fake cheese
        - locked: cheese move only when unlock (when player collect more than 'unlock_score' score)
        """
        pygame.sprite.Sprite.__init__(self)
        self.game_state = game_state
        self.speed = gs.CHEESE_SPEED
        self.typ = typ               # 0,1=normal / 2=bonus /  3,4=fake / 5=bonus fake
        self.fake = False            # Determines if the cheese is FAKE type = 3, 4, 5
        self.add_to_stats = False    # Add to TOTAL_CHEESE/FAKE when shown on the screen
        self.extra_distance = random.randint(gs.CHEESE_min_distance, gs.CHEESE_max_distance)
        self.bonus_chance = bonus_chance
        self.fake_chance = fake_chance
        self.locked = True if unlock_score > 0 else False
        self.was_locked = True if unlock_score > 0 else False
        self.unlock_score = unlock_score
        self.score = random.randint(11, 15)

        self.image_list = images
        self.image = images[typ]
        self.rect = self.image.get_rect(center=(x, y))

    def update(self):
        """ move and restart after not catch """
        if not self.locked:
            self.rect.y += self.speed

        if not self.add_to_stats and self.rect.top >= gs.CHEESE_offset:
            self.add_to_stats = True

        if self.rect.top > gs.HUD2:
            if self.fake:
                self.game_state.FAKE_IN_ROW = 0
            else:
                self.game_state.MISS_IN_ROW += 1
            self.restart()

    def draw(self, surface):
        """ draw images if not locked """
        if not self.locked:
            surface.blit(self.image, self.rect)

    def restart(self):
        """ reappears after being caught or running away from the game screen """
        if self.fake and self.add_to_stats:
            self.game_state.TOTAL_FAKE += 1
        elif not self.fake and self.add_to_stats:
            self.game_state.TOTAL_CHEESE += 1

        # restart type
        self.typ = 2 if self.bonus_chance >= random.randint(0, 100) else random.randint(0, 1)
        if self.fake_chance >= random.randint(0, 100):
            self.typ += 3
            self.fake = True
        else:
            self.fake = False

        # restart settings
        self.add_to_stats = False
        self.image = self.image_list[self.typ]
        self.rect.width, self.rect.height = self.image.get_size()
        self.score = random.randint(23, 28) if self.typ == 2 else random.randint(11, 15)

        # reduce respawn distance (time)
        self.extra_distance = gs.random_num(a=gs.CHEESE_min_distance,
                                            b=gs.CHEESE_max_distance - gs.CHEESE_distance_reduction,
                                            to_difference=False)
        # restart position
        while True:
            # checks that cheeses don't respawn next to each other
            self.rect.x = gs.random_num(gs.CHEESE_offset, gs.WINDOW_WIDTH)
            self.rect.y = -self.extra_distance
            too_close = any(
                gs.diagonal_distance2(pos[0], pos[1], self.rect.x, self.rect.y) < gs.CHEESE_distance_respawn
                for pos in self.game_state.SPAWN_POS_LIST)

            # update SPAWN_POS_LIST and cheese position
            if not too_close:
                new_spawn_pos = (self.rect.x, self.rect.y)
                self.game_state.SPAWN_POS_LIST.append(new_spawn_pos)
                self.game_state.SPAWN_POS_LIST = self.game_state.SPAWN_POS_LIST[-self.game_state.LIST_SIZE:]
                break

    def default_settings(self):
        """ restores the initial settings (loads faster instead of deleting and adding to the group) """
        self.speed = gs.CHEESE_SPEED
        self.extra_distance = gs.CHEESE_max_distance
        self.locked = self.was_locked
        self.restart()


class Drop(pygame.sprite.Sprite):
    def __init__(self, typ: str, images: list, animation_images: list = None, speed: int = 0, x: int = 0, y: int = 0):
        """
         - typ: water / poison / egg / cola / time
         - animation_img_list - additional list of images to create an additional animation
                                (replaces the image from the image_list)
        """
        pygame.sprite.Sprite.__init__(self)
        # settings
        self.data = gs.ITEM_dict.get(typ, 0)
        self.typ = typ
        self.speed = speed
        self.basic_speed = speed
        self.freeze = False            # True = object dont move

        # basic animation
        self.update_time = pygame.time.get_ticks()
        self.image_index = 0
        self.image_list = images
        self.basic_image_list = images
        self.image = self.image_list[self.image_index]
        self.rect = self.image.get_rect(center=(x, y))

        # extra animation
        self.animation = False
        self.animation_list = animation_images
        self.animation_cooldown = self.data[0]
        self.extra_time = self.data[1]
        self.crush_distance = 100000
        if self.data[10] > 0:
            self.crush_distance = random.randint(gs.MAX_HEIGHT - self.data[10], gs.MAX_HEIGHT)

        # rect for collision
        self.rect_pos_changed = False
        self.double_resize_rect = False
        self.rect_col = self.image.get_rect(center=(x, y))
        self.rect_first_resize = self.data[2]
        self.rect_second_resize = self.data[3]

        # sounds
        self.sounds_played = False
        self.sound_collision = gs.load_sound(self.data[11])
        self.sound_animation = gs.load_sound(self.data[12])

    def update(self, player: Player):
        # update image
        self.image = self.image_list[self.image_index]
        self.image.set_colorkey(gs.BLACK)

        # update collision rect
        self.rect.width, self.rect.height = self.image.get_size()
        if not self.double_resize_rect:
            self.rect_col = gs.resize_and_offset_rect(self.rect, self.rect_first_resize)
        else:
            self.rect_col = gs.resize_and_offset_rect(gs.resize_and_offset_rect(self.rect, self.rect_first_resize),
                                                      self.rect_second_resize)

        # make basic animation - hourglass
        if not self.animation and pygame.time.get_ticks() - self.update_time > self.animation_cooldown:
            if len(self.image_list) >= 2:             # don't make animations for dropped Eggs / Water
                self.image_index = (self.image_index + 1) % len(self.image_list)
                self.update_time = pygame.time.get_ticks()

        # make animation - egg or water crush
        if self.animation and pygame.time.get_ticks() - self.update_time > self.animation_cooldown:
            if self.image_index == len(self.image_list) - 1:
                self.double_resize_rect = True
                if pygame.time.get_ticks() - self.update_time > self.animation_cooldown + self.extra_time:
                    self.restart()
            else:
                self.image_index += 1
                self.update_time = pygame.time.get_ticks()

        # move - update rect
        self.rect.y += self.speed if not self.freeze else 0

        # restart after passing through the wall
        if self.rect.top > gs.HUD2:
            self.restart()

        # make crush animation
        if self.crush_distance <= self.rect.bottom and not self.animation:
            self.make_animation()
            self.sound_animation.play()

        # check collision with player
        if self.rect_col.colliderect(player.rect) and not player.dead:
            if self.typ == 'time':
                player.score += 1
                player.time_left += 15
                self.restart()
                self.sound_collision.play()
            if self.typ == 'cola':
                if player.time_left > 15:
                    player.score += 25
                    player.update_speed(random.randint(7, 9) / 100)
                    player.time_left -= 9.75
                    player.time_left *= 0.90
                else:
                    player.score -= 30
                self.restart()
                self.sound_collision.play()
            elif self.typ == 'egg':
                if self.animation:
                    player.health -= 0.35
                else:
                    player.health -= 11.5
                    self.sound_collision.play()
                    self.make_animation()
            elif self.typ == 'water':
                player.boost_lvl += (player.max_boost - player.boost_lvl) * 0.5
                player.time_left += 0.50
                player.penetration = True
                player.wraps += 1
                self.sound_collision.play()
                self.restart()
            elif self.typ == 'poison':
                player.boost_lvl -= 250
                player.poison += random.randint(11, 15)
                player.slowed = True
                player.slow_time = pygame.time.get_ticks()
                player.update_speed(-0.02)
                self.sound_collision.play()
                self.restart()

    def draw(self, surface):
        surface.blit(self.image, self.rect)

    def restart(self):
        # restart of coordinates
        self.rect.x = gs.random_num(self.data[4], gs.WINDOW_WIDTH)
        self.rect.y = -random.randint(self.data[8], self.data[9])

        # restart setting
        self.double_resize_rect = False
        self.rect_pos_changed = False
        self.sounds_played = False
        self.animation = False
        self.freeze = False
        # self.restart_sounds(volume=0)
        self.image_index = 0
        self.image_list = self.basic_image_list
        self.crush_distance = 100000
        if self.data[10] > 0:
            self.crush_distance = random.randint(gs.MAX_HEIGHT - self.data[10], gs.MAX_HEIGHT)

    def restart_speed(self):
        self.speed = self.basic_speed

    def restart_sounds(self, volume):
        if self.sound_animation:
            self.sound_animation.set_volume(volume)
            self.sound_animation.play()
            self.sound_animation.stop()

    def make_animation(self):
        prev_img_width = self.image.get_width()
        prev_img_height = self.image.get_height()
        self.animation = True
        self.freeze = True
        self.update_time = pygame.time.get_ticks()
        self.image_index = 0
        self.image_list = self.animation_list

        # Graphics sizes vary, animation position adjustment
        self.rect.x -= self.image.get_width() // 2 + prev_img_width
        self.rect.y -= self.image.get_height() // 2 + prev_img_height


class Trap(pygame.sprite.Sprite):
    def __init__(self, game_state: GameState, x: int, y: int, images: list, unlock_score: int = 0,
                 sounds_start: pygame.mixer.Sound = None, sound_hit: pygame.mixer.Sound = None):
        pygame.sprite.Sprite.__init__(self)
        self.game_state = game_state
        self.pause = True
        self.animation = False
        self.sound = None
        self.locked = True if unlock_score > 0 else False
        self.was_locked = True if unlock_score > 0 else False
        self.unlock_score = unlock_score
        self.typ = "mousetrap"

        # sounds
        self.sound_start = sounds_start
        self.sound_hit = sound_hit
        self.sound_play = True

        # images
        self.image_index = 2
        self.images_list = images
        self.image = self.images_list[self.image_index]
        self.rect = self.image.get_rect(center=(x, y))

        # collision rects
        self.collision_rects = [
            pygame.Rect(x + 24, y + 86, 60, 19),
            pygame.Rect(x + 32, y + 67, 60, 19),
            pygame.Rect(x + 40, y + 48, 60, 19),
            pygame.Rect(x + 60, y + 28, 50, 20)]

        # time settings
        self.animation_cooldown = 80
        self.draw_cooldown = gs.TRAP_draw_time
        self.pause_cooldown = random.randint(gs.TRAP_pause_min, gs.TRAP_pause_max)
        self.update_animation_time = pygame.time.get_ticks()
        self.update_draw_time = pygame.time.get_ticks()
        self.update_pause_time = pygame.time.get_ticks()
        self.start_pause()

    def update(self, player: Player):
        # unlock trap
        if self.locked:
            if player.score >= self.unlock_score:
                self.locked = False

        else:
            self.image = self.images_list[self.image_index]
            if self.animation:
                if self.sound_play:
                    self.sound_start.play()
                    self.sound_play = False
                # make explosion animation
                if pygame.time.get_ticks() - self.update_animation_time > self.animation_cooldown:
                    if self.image_index == len(self.images_list) - 1:
                        if pygame.time.get_ticks() - self.update_animation_time >= self.animation_cooldown - 5:
                            self.image_index = 0
                            self.stop_animation()
                    else:
                        self.image_index += 1
                        self.update_animation_time = pygame.time.get_ticks()
            else:
                # Check collision with player
                for rect in self.collision_rects:
                    if rect.colliderect(player.rect) and not player.dead and not self.pause:
                        if self.typ == "mousetrap":
                            player.slowed = True
                            player.slow_time = pygame.time.get_ticks()
                            player.health -= random.randint(25, 33)
                        elif self.typ == "medicine kit":
                            if player.time_left >= 20:
                                player.health += int((player.max_hp - player.health) * 0.15)
                                player.time_left -= 15.75
                                player.score += 25
                            else:
                                player.health += 1
                                player.score -= 30
                        self.sound_hit.play()
                        self.start_pause()

                # draw mousetrap
                if not self.pause and pygame.time.get_ticks() - self.update_draw_time > self.draw_cooldown:
                    self.start_pause()

                # remove mousetrap
                if self.pause and pygame.time.get_ticks() - self.update_pause_time > self.pause_cooldown:
                    self.stop_pause()

    def draw(self, surface):
        if not self.locked:
            surface.blit(self.image, self.rect)

    def start_pause(self):
        self.pause = True
        self.sound_play = True
        if self.typ == "medicine kit":
            self.pause_cooldown = random.randint(gs.MEDICINE_pause_min, gs.MEDICINE_pause_max)
        else:
            self.pause_cooldown = random.randint(gs.TRAP_pause_min, gs.TRAP_pause_max)
        self.update_pause_time = pygame.time.get_ticks()
        self.rect.x = -200
        self.rect.y = 1000

    def stop_pause(self):
        self.pause = False
        self.make_animation()
        while True:
            # respawn in random pos.
            self.rect.x = random.randint(150, 1050)
            self.rect.y = random.randint(100, 500)

            # update collision rect
            if self.typ == "medicine kit":
                self.collision_rects = [pygame.Rect(self.rect.x + 30, self.rect.y + 40, 75, 55)]
            else:
                self.collision_rects = [
                    pygame.Rect(self.rect.x + 24, self.rect.y + 86, 60, 19),
                    pygame.Rect(self.rect.x + 32, self.rect.y + 67, 60, 19),
                    pygame.Rect(self.rect.x + 40, self.rect.y + 48, 60, 19),
                    pygame.Rect(self.rect.x + 60, self.rect.y + 28, 50, 20)]

            # checks for no respawn next to each other
            too_close = any(
                gs.diagonal_distance2(pos[0], pos[1], self.rect.x, self.rect.y) < 200
                for pos in self.game_state.TRAP_POS_LIST)

            # update TRAP_POS_LIST and cheese position
            if not too_close:
                trap_spawn_pos = (self.rect.x, self.rect.y)
                self.game_state.TRAP_POS_LIST.append(trap_spawn_pos)
                self.game_state.TRAP_POS_LIST = self.game_state.TRAP_POS_LIST[-3:]
                break

    def make_animation(self):
        if not self.animation:
            self.pause = False
            self.animation = True
            self.update_animation_time = pygame.time.get_ticks()
            self.image_index = 2

    def stop_animation(self):
        self.animation = False
        self.image_index = 1
        self.update_draw_time = pygame.time.get_ticks()

    def restart(self):
        self.locked = self.was_locked
        self.rect.x = -500
        self.rect.y = -500
# endregion


# region 5. Create Game Data

# Create Object
game_state = GameState()
player = Player(game_state=game_state, images=img_list_player, x=500, y=600, typ=1)
player2 = Player(game_state=game_state, images=img_list_player, x=700, y=600, typ=2)

cola = Drop(typ='cola', images=img_list_soda, speed=3.75, x=gs.random_num(25, 1200), y=-250)
egg = Drop(typ='egg', images=[image_egg], animation_images=img_list_egg, speed=3.65,
           x=gs.random_num(25, 1200), y=gs.random_num(50, -300))
water = Drop(typ='water', images=[image_water], animation_images=img_list_water, speed=3.85,
             x=gs.random_num(25, 1200), y=gs.random_num(50, -200))
poison = Drop(typ='poison', images=[image_poison], animation_images=img_list_poison, speed=3.80,
              x=gs.random_num(25, 1200), y=gs.random_num(50, -500))
medicine_kit = Trap(game_state=game_state, x=730, y=120, images=img_list_medicineKit, sounds_start=sound_trap_start,
                    sound_hit=sound_medicine_hit, unlock_score=1250)
medicine_kit.typ = "medicine kit"
medicine_kit.draw_cooldown = gs.MEDICINE_draw_time
medicine_kit.pause_cooldown = 25000

mousetrap_01 = Trap(game_state=game_state, x=330, y=220, images=img_list_mousetrap, sounds_start=sound_trap_start,
                    sound_hit=sound_trap_hit)
mousetrap_02 = Trap(game_state=game_state, x=630, y=220, images=img_list_mousetrap, unlock_score=550,
                    sounds_start=sound_trap_start, sound_hit=sound_trap_hit)
mousetrap_03 = Trap(game_state=game_state, x=930, y=220, images=img_list_mousetrap, unlock_score=1350,
                    sounds_start=sound_trap_start, sound_hit=sound_trap_hit)
mousetrap_04 = Trap(game_state=game_state, x=230, y=360, images=img_list_mousetrap, unlock_score=2000,
                    sounds_start=sound_trap_start, sound_hit=sound_trap_hit)

# Create cheese group
cheese_group = pygame.sprite.Group()
for cheese_index in range(1, 5):
    cheese_01 = Cheese(game_state=game_state, x=240 * cheese_index, y=-random.randint(50, 350),
                       images=img_list_cheese, bonus_chance=gs.CHEESE_bonus_chance, fake_chance=gs.CHESSE_fake_chance)
    cheese_group.add(cheese_01)

cheese_02 = Cheese(game_state=game_state, x=gs.random_num(gs.CHEESE_offset, gs.WINDOW_WIDTH), y=-150,
                   images=img_list_cheese, bonus_chance=10, fake_chance=20, unlock_score=450)
cheese_group.add(cheese_02)

cheese_03 = Cheese(game_state=game_state, x=gs.random_num(gs.CHEESE_offset, gs.WINDOW_WIDTH), y=-150,
                   images=img_list_cheese, bonus_chance=20, fake_chance=35, unlock_score=1250)
cheese_group.add(cheese_03)

cheese_04 = Cheese(game_state=game_state, x=gs.random_num(gs.CHEESE_offset, gs.WINDOW_WIDTH), y=-150,
                   images=img_list_cheese, bonus_chance=25, fake_chance=50, unlock_score=1750)
cheese_group.add(cheese_04)

cheese_05 = Cheese(game_state=game_state, x=gs.random_num(gs.CHEESE_offset, gs.WINDOW_WIDTH), y=-150,
                   images=img_list_cheese, bonus_chance=25, fake_chance=0, unlock_score=2250)
cheese_group.add(cheese_05)

# Create items group
item_group = pygame.sprite.Group()
item_group.add(cola)
item_group.add(egg)
item_group.add(water)
item_group.add(poison)

# create trap group
trap_group = pygame.sprite.Group()
trap_group.add(medicine_kit)
trap_group.add(mousetrap_01)
trap_group.add(mousetrap_02)
trap_group.add(mousetrap_03)
trap_group.add(mousetrap_04)

# Create MAIN group
mutter_group = pygame.sprite.Group()
mutter_group.add(cheese_group)
mutter_group.add(item_group)
mutter_group.add(trap_group)


def restart_game(r_time=750):
    global loading_screen
    global money

    # Clear display
    loading_screen = True
    if loading_screen:
        WG.fill(color=gs.BLACK)
        WG.blit(image_loading, rect_center)
        pygame.display.flip()

    # Restart items
    for r_item in item_group:
        r_item.restart()
        r_item.restart_speed()

    # Restart trap
    for t_item in trap_group:
        t_item.restart()
    medicine_kit.pause_cooldown = 25000

    # Restart cheese
    game_state.restart_cheese()
    for r_cheese in cheese_group:
        r_cheese.default_settings()

    # Add Money
    money += max(0, player.score // 15 + 2 * player.cheese_amount - 125)
    money += max(0, player2.score // 15 + 2 * player2.cheese_amount - 125)

    # Restart Players
    player.restart(img_index=player1_skin_image_index)
    if number_of_players == 2:
        player2.restart(img_index=player2_skin_image_index)

    # Restart screen
    if r_time >= 50:
        pygame.time.delay(r_time)
    loading_screen = False
# endregion


# region 6.0. Create Buttons - Main Menu
# basic
button_x = gs.WINDOW_WIDTH // 2 - 450
button_y = 200
button_size = 100

button_00_00 = ButtonText(x=button_x, y=button_y + 0 * button_size, width=450, height=button_size - 25,
                          text="Single Player", font=gs.FONT_65, default_color=menu_color, hover_color=select_color)
button_00_01 = ButtonText(x=button_x, y=button_y + 1 * button_size, width=450, height=button_size - 25,
                          text="Multi Player", font=gs.FONT_65, default_color=menu_color, hover_color=select_color)
button_00_02 = ButtonText(x=button_x, y=button_y + 2 * button_size, width=450, height=button_size - 25,
                          text="High Score", font=gs.FONT_65, default_color=menu_color, hover_color=select_color)
button_00_03 = ButtonText(x=button_x, y=button_y + 3 * button_size, width=450, height=button_size - 25,
                          text="Settings", font=gs.FONT_65, default_color=menu_color, hover_color=select_color)
button_00_04 = ButtonText(x=button_x, y=button_y + 4 * button_size, width=450, height=button_size - 25,
                          text="Quit Game", font=gs.FONT_65, default_color=menu_color, hover_color=select_color)

# extra
hover_color = gs.PINK_MENU
button_00_10 = ButtonRect(top_left=(637, 195), bottom_right=(722, 280), default_color=gs.BLACK,
                          hover_color=hover_color, time=600, description="Restart Game")
button_00_11 = ButtonRect(top_left=(637, 283), bottom_right=(722, 363), default_color=gs.BLACK,
                          hover_color=hover_color, time=600, description="How To Play?")
button_00_12 = ButtonRect(top_left=(637, 372), bottom_right=(722, 452), default_color=gs.BLACK,
                          hover_color=hover_color, time=600, description="View sources")
button_00_13 = ButtonRect(top_left=(637, 461), bottom_right=(722, 542), default_color=gs.BLACK,
                          hover_color=hover_color, time=600, description="Change Controls")
button_00_14 = ButtonRect(top_left=(637, 545), bottom_right=(722, 630), default_color=gs.BLACK,
                          hover_color=hover_color, time=600, description="Open Store")
buttons_start = [button_00_10, button_00_11, button_00_12, button_00_13, button_00_14]
# endregion

# region 6.1. Create Buttons - High Score Menu
# basic
button_01_back = ButtonRect(top_left=(830, 545), bottom_right=(950, 695), default_color=gs.WHITE, hover_color=gs.BLUE)
button_01_restart_TP = ButtonRect(top_left=(705, 600), bottom_right=(805, 680), default_color=gs.WHITE,
                                  hover_color=gs.BLUE, time=500, action=True)
button_01_quit = ButtonRect(top_left=(985, 640), bottom_right=(1080, 710), default_color=gs.WHITE, hover_color=gs.BLUE)

# vote restart list
button_01_vote_yes = ButtonText(x=gs.WINDOW_WIDTH - 430, y=90, width=125, height=70, frame=False,
                                text="yes", font=gs.FONT_60, default_color=gs.GREEN, hover_color=gs.YELLOW_2)
button_01_vote_no = ButtonText(x=gs.WINDOW_WIDTH - 280, y=90, width=125, height=70, frame=False,
                               text="NO", font=gs.FONT_60, default_color=gs.RED, hover_color=gs.YELLOW_2)

# extra
hover_color = gs.BLUE
button_01_10 = ButtonRect(top_left=(555, 45), bottom_right=(640, 120), default_color=gs.BLACK,
                          hover_color=hover_color, time=45, description="Frenzy Color")
button_01_11 = ButtonRect(top_left=(555, 120), bottom_right=(635, 210), default_color=gs.BLACK,
                          hover_color=hover_color, time=45, description="Solid Color")
button_01_12 = ButtonRect(top_left=(555, 210), bottom_right=(640, 295), default_color=gs.BLACK,
                          hover_color=hover_color, time=60, description="Change to Top 10 List")
button_01_13 = ButtonRect(top_left=(555, 295), bottom_right=(635, 385), default_color=gs.BLACK,
                          hover_color=hover_color, time=600, description="Save & Open Global Ranking")
buttons_score = [button_01_10, button_01_11, button_01_12, button_01_13]
# endregion

# region 6.2 Create Buttons - Settings Menu
# basic
button_02_back = ButtonRect(top_left=(680, 615), bottom_right=(795, 735), default_color=gs.BLACK, hover_color=gs.YELLOW)
button_02_how2play = ButtonRect(top_left=(805, 615), bottom_right=(920, 735), default_color=gs.BLACK,
                                hover_color=gs.YELLOW, time=125)
button_02_shop = ButtonRect(top_left=(930, 615), bottom_right=(1050, 735), default_color=gs.BLACK,
                            hover_color=gs.YELLOW, time=125)
button_02_sources = ButtonRect(top_left=(1060, 615), bottom_right=(1180, 735), default_color=gs.BLACK,
                               hover_color=gs.YELLOW, time=300)

# players images change index
change_cooldown = 15
button_02_p1_prev = ButtonRect(top_left=(915, 160), bottom_right=(955, 200), default_color=gs.BLACK,
                               hover_color=gs.YELLOW, time=change_cooldown)
button_02_p1_next = ButtonRect(top_left=(1085, 160), bottom_right=(1130, 200), default_color=gs.BLACK,
                               hover_color=gs.YELLOW, time=change_cooldown)
button_02_p2_prev = ButtonRect(top_left=(915, 350), bottom_right=(955, 390), default_color=gs.BLACK,
                               hover_color=gs.YELLOW, time=change_cooldown)
button_02_p2_next = ButtonRect(top_left=(1085, 350), bottom_right=(1130, 390), default_color=gs.BLACK,
                               hover_color=gs.YELLOW, time=change_cooldown)
button_02_random = ButtonRect(top_left=(1085, 225), bottom_right=(1130, 265), default_color=gs.BLACK,
                              hover_color=gs.YELLOW, time=change_cooldown * 2)

# settings options
button_02_restore_TP = ButtonRect(top_left=(520, 300), bottom_right=(605, 385), default_color=gs.BLACK,
                                  hover_color=gs.YELLOW, time=180000)
button_02_remove_TP = ButtonRect(top_left=(605, 300), bottom_right=(695, 385), default_color=gs.BLACK,
                                 hover_color=gs.YELLOW, time=60000)
button_02_controls = ButtonRect(top_left=(695, 300), bottom_right=(780, 385), default_color=gs.BLACK,
                                hover_color=gs.YELLOW, time=10000)
button_02_music = ButtonRect(top_left=(520, 395), bottom_right=(605, 480), default_color=gs.BLACK,
                             hover_color=gs.PINK, time=5)
button_02_buy_skin = ButtonRect(top_left=(920, 370), bottom_right=(1005, 450), default_color=gs.BLACK,
                                hover_color=gs.YELLOW, time=1000)
button_02_bank = ButtonRect(top_left=(920, 288), bottom_right=(1005, 368), default_color=gs.BLACK,
                            hover_color=gs.YELLOW, time=5)
# endregion

# region 6.3 Create Buttons - Control Menu
change_cooldown = 50
button_03_back = ButtonRect(top_left=(575, 575), bottom_right=(680, 675), default_color=gs.BLACK, hover_color=gs.BLUE)
button_03_disk_prev = ButtonRect(top_left=(780, 685), bottom_right=(830, 725), default_color=gs.BLACK,
                                 hover_color=gs.YELLOW, time=change_cooldown)
button_03_disk_next = ButtonRect(top_left=(955, 685), bottom_right=(995, 725), default_color=gs.BLACK,
                                 hover_color=gs.YELLOW, time=change_cooldown)

button_03_yellow = ButtonRect(top_left=(710, 575), bottom_right=(815, 675), default_color=gs.BLACK, hover_color=gs.BLUE)
button_03_pink = ButtonRect(top_left=(845, 575), bottom_right=(950, 675), default_color=gs.BLACK, hover_color=gs.BLUE)
button_03_blue = ButtonRect(top_left=(980, 575), bottom_right=(1085, 675), default_color=gs.BLACK, hover_color=gs.BLUE)
# endregion

# region 6.4 Create Buttons - Info Menu
button_04_prev = ButtonRect(top_left=(765, 180), bottom_right=(885, 300), default_color=gs.BLACK,
                            hover_color=gs.YELLOW, time=75)
button_04_next = ButtonRect(top_left=(900, 180), bottom_right=(1020, 300), default_color=gs.BLACK,
                            hover_color=gs.YELLOW, time=75)
button_04_back = ButtonRect(top_left=(1030, 180), bottom_right=(1150, 300), default_color=gs.BLACK,
                            hover_color=gs.YELLOW, time=2500)
# endregion

# region 6.5. Create Buttons - Game Over / Restart Game Screen
button_05_00 = ButtonText(x=70, y=650, width=275, height=65,
                          text="Main Menu", font=gs.FONT_50, default_color=gs.DARK_WHITE, hover_color=gs.YELLOW)
button_05_01 = ButtonText(x=400, y=650, width=350, height=65,
                          text="Restart Game", font=gs.FONT_50, default_color=gs.WHITE, hover_color=gs.YELLOW)
button_05_02 = ButtonText(x=825, y=650, width=275, height=65,
                          text="Quit Game", font=gs.FONT_50, default_color=gs.DARK_WHITE, hover_color=gs.YELLOW)
# endregion

# region 6.6. Create Buttons - Licenses Menu
button_06_00 = ButtonRect(top_left=(125, 430), bottom_right=(280, 585), default_color=gs.BLACK, hover_color=gs.YELLOW,
                          description="Back to previous menu")
button_06_01 = ButtonRect(top_left=(290, 430), bottom_right=(450, 585), default_color=gs.BLACK, hover_color=gs.YELLOW,
                          description="Open in web browser")
button_06_02 = ButtonRect(top_left=(455, 430), bottom_right=(605, 585), default_color=gs.BLACK, hover_color=gs.YELLOW,
                          description="Open all in Excel", time=1000)

button_06_prev = ButtonRect(top_left=(285, 350), bottom_right=(360, 415), default_color=gs.BLACK,
                            hover_color=gs.YELLOW, time=15, description="Previous object")
button_06_next = ButtonRect(top_left=(375, 350), bottom_right=(445, 415), default_color=gs.BLACK,
                            hover_color=gs.YELLOW, time=15, description="Next object")
# endregion


# region 7. Start Game
intro_screen = False  # Stop Intro window

while game_running:
    for event in pygame.event.get():
        # Exit Game
        if event.type == pygame.QUIT:
            game_running = False
        elif event.type == pygame.KEYDOWN:
            # Pause Game
            if event.key == pygame.K_p:
                game_paused = not game_paused
                music_player()
                pause_game_image_index = random.randint(0, len(img_list_pause_game) - 1)
            # Back to MainMenu
            elif event.key == pygame.K_ESCAPE:
                if game_paused:
                    game_over = True
                elif game_over:
                    restart_game(r_time=0)
                    game_screen = False
                    menu_screen = True
                    menu_start = True
                elif menu_screen:
                    if menu_licenses:
                        menu_licenses = False
                        menu_start = False
                        menu_settings = True
                    elif menu_start:
                        game_running = False
                    elif menu_controls:
                        if editing_key is None:
                            menu_licenses = False
                            menu_start = False
                            menu_settings = True
                    else:
                        menu_start = True
                        menu_score = False
                        menu_settings = False
            # Restart game after GameOver
            elif event.key == pygame.K_RETURN:
                if game_over and number_of_players == 1:
                    game_over = False
                    game_paused = False
                    player.restart(img_index=player1_skin_image_index)
                    restart_game()
            # Change controls
            if menu_controls:
                # Player 1
                if editing_key is not None:
                    controls[editing_key] = event.key
                    editing_key = None
                else:
                    for action, key in controls.items():
                        if event.key == key:
                            editing_key = action

    # Mouse Position (debug)
    mouse_x, mouse_y = pygame.mouse.get_pos()
    if debug_mouse_pos:
        mouse_counter += 1
        if mouse_counter >= 60:
            mouse_counter = 0
            print(f"X: {mouse_x}, Y: {mouse_y}")

    if menu_screen:
        # Open Main Menu
        if menu_start:
            WG.fill(color=gs.BLACK)

            # Draw Buttons - basic
            if button_00_00.draw(surface=WG):  # Button - Single Player
                if skin_not_owned:  # restart skins if not owned
                    player1_skin_image_index = 0
                    player2_skin_image_index = 1
                    player.update_image(img_index=player1_skin_image_index)
                    player2.update_image(img_index=player2_skin_image_index)
                menu_screen = False
                menu_start = False
                game_screen = True
                number_of_players = 1
                last_update_time = time.time()
                intro_sounds.stop()

            elif button_00_01.draw(surface=WG):  # Button - Two Player
                if skin_not_owned:  # restart skins if not owned
                    player1_skin_image_index = 0
                    player2_skin_image_index = 1
                    player.update_image(img_index=player1_skin_image_index)
                    player2.update_image(img_index=player2_skin_image_index)
                menu_screen = False
                menu_start = False
                game_screen = True
                number_of_players = 2
                last_update_time = time.time()
                intro_sounds.stop()

            elif button_00_02.draw(surface=WG):  # Button - High Score
                menu_score = True
                menu_start = False

            elif button_00_03.draw(surface=WG):  # Button - Settings
                menu_settings = True
                menu_start = False

            elif button_00_04.draw(surface=WG):  # Button - Quit Game
                game_running = False

            # Draw Buttons - extra
            start_text = ""
            for button in buttons_start:
                if button.draw(surface=WG):
                    if button == button_00_10:  # restart game
                        restart_game(r_time=150)
                    elif button == button_00_11:  # how to play
                        menu_info = True
                        menu_start = False
                    elif button == button_00_12:  # licenses menu
                        menu_licenses = True
                        menu_start = False
                    elif button == button_00_13:  # control menu
                        menu_controls = True
                        menu_start = False
                    elif button == button_00_14:  # shop menu
                        menu_shop = True
                        menu_settings = True
                        menu_start = False

                if button.mouse_in():
                    start_text = button.description

            # Draw images
            gs.draw_image(surface=WG, image=image_menu_start, x=gs.WINDOW_WIDTH - 450, y=gs.WINDOW_HEIGHT - 450)
            gs.draw_image(surface=WG, image=image_buttons_start_menu, x=625, y=175)

            # Draw text
            gs.draw_text(surface=WG, text="Feed The Mouse", font=gs.FONT_110, color=gs.YELLOW, x=550, y=110)
            gs.draw_text(surface=WG, text="3", font=gs.FONT_120, color=gs.WHITE, x=1055, y=115)
            gs.draw_text(surface=WG, text=start_text, x=960, y=225, font=gs.FONT_36)
            if start_text:
                gs.draw_image(surface=WG, image=image_info, x=900 - gs.text_width(start_text, gs.FONT_36) // 2, y=200)

        # Open HighScore Menu
        if menu_score:
            WG.fill(color=gs.BLACK)

            # Draw Basic Buttons
            if button_01_back.draw(WG):  # Back Main Menu
                menu_score = False
                menu_start = True
            if button_01_restart_TP.draw(WG):  # Restart ScoreList
                pass
            if button_01_quit.draw(WG):  # Quit Game
                game_running = False

            # Draw Extra Buttons
            score_text = ""
            for button in buttons_score:
                if button.draw(surface=WG):
                    if button == button_01_10:
                        COLORS = gs.create_colors_list(size=30, remove_gray=True, only_contrast=True,
                                                       only_one_color=False)
                    elif button == button_01_11:
                        high_score_color_index = (high_score_color_index + 1) % len(high_score_color_list)
                        if high_score_color_index == 0 and high_score_image_index == 1:
                            high_score_color_index += 1
                        COLORS = gs.create_colors_list(size=30, remove_gray=False, only_contrast=False,
                                                       only_one_color=True,
                                                       color=high_score_color_list[high_score_color_index])
                    elif button == button_01_12:
                        if high_score_image_index == 2:
                            high_score_image_index = 1
                            COLORS = gs.create_colors_list(size=30, remove_gray=True, only_contrast=True,
                                                           only_one_color=False)
                        elif high_score_image_index == 1:
                            high_score_image_index = 2
                            high_score_color_index = 0
                            COLORS = gs.create_colors_list(size=30, remove_gray=False, only_contrast=False,
                                                           only_one_color=True,
                                                           color=high_score_color_list[high_score_color_index])
                    elif button == button_01_13:
                        webbrowser.open("discord.gg/um6zuuKhxG")
                        print("open DC")

                if button.mouse_in():
                    score_text = button.description
                    if button == button_01_12 and high_score_image_index == 1:
                        score_text = "Change to Top 30 List"

            # Draw Images
            gs.draw_image(surface=WG, image=img_list_high_score[high_score_image_index], x=50, y=0)
            gs.draw_image(surface=WG, image=img_list_high_score[0], x=gs.WINDOW_WIDTH - 575, y=gs.WINDOW_HEIGHT - 600)
            gs.draw_image(surface=WG, image=image_buttons_score_menu, x=545, y=30)

            # Draw Text
            gs.draw_text(surface=WG, text=score_text, x=910, y=100, font=gs.FONT_36)
            if score_text:
                gs.draw_image(surface=WG, image=image_info, x=850 - gs.text_width(score_text, gs.FONT_36) // 2, y=75)

            # Load ScoreList Data
            if load_high_scores:
                wb = openpyxl.load_workbook(excel_file_path)
                ws = wb.active
                sorted_results = sorted(ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row),
                                        key=lambda var_x: var_x[0].value, reverse=True)
                top10_score_list = sorted_results[:10]  # top 10 best score
                top30_score_list = sorted_results[:30]  # top 30 best score

            # Draw Best Score
            if high_score_image_index == 1:
                for i, result in enumerate(top10_score_list):
                    text_y = 70 if i == 0 else 155 + 63 * (i - 1) + 21 * (i // 5)
                    gs.draw_text(surface=WG, text=f"{i + 1}.", x=130, y=text_y, font=gs.FONT_65, color=COLORS[i])
                    gs.draw_text(surface=WG, text=f"{result[0].value}", x=280, y=text_y,
                                 font=gs.FONT_65, color=COLORS[i + 10])
                    gs.draw_text(surface=WG, text=f"{result[2].value}", x=450, y=text_y,
                                 font=gs.FONT_65, color=COLORS[i + 20])
                    if i > 0:
                        gs.draw_text(surface=WG, text="-", x=180, y=text_y, font=gs.FONT_65, color=gs.WHITE)
                        gs.draw_text(surface=WG, text="-", x=370, y=text_y, font=gs.FONT_65, color=gs.WHITE)
            elif high_score_image_index == 2:
                for i, result in enumerate(top30_score_list):
                    score_text = f"{result[0].value}-{result[2].value}"
                    if 0 <= i <= 2:
                        text_y = 215 + 65 * i - max(0, i - 1) * 10
                        gs.draw_text(surface=WG, text=score_text, x=365, y=text_y, font=gs.FONT_60, color=COLORS[i])
                    else:
                        text_y = 385 + 37 * (i - 3) - 37 * 9 * ((i - 3) // 9)
                        text_x = 110 + 135 * ((i - 3) // 9)
                        gs.draw_text(surface=WG, text=f"{i + 1}.", x=text_x, y=text_y, font=gs.FONT_25, color=COLORS[i])
                        gs.draw_text(surface=WG, text=score_text, x=text_x + 60, y=text_y, font=gs.FONT_25,
                                     color=COLORS[i])

            # Restart score
            if button_01_restart_TP.freeze:
                gs.draw_text(surface=WG, text='Restart Score List?', font=gs.FONT_50, color=gs.YELLOW_2, x=885, y=55)

                if button_01_vote_yes.draw(WG):  # Restart? YES
                    button_01_restart_TP.restart()
                    gs.rename_file(old_name='score', new_name='score_backUp')
                if button_01_vote_no.draw(WG):  # Restart? NO
                    button_01_restart_TP.restart()

        # Open Settings Menu
        if menu_settings:
            WG.fill(color=gs.BLACK)
            skin_not_owned = False
            purchase_text = " "

            # Draw music and bank cash slider bar
            pygame.draw.rect(WG, gs.PINK_MUSIC, (587, 400, 10 + volume * 2, 55))
            pygame.draw.circle(WG, gs.BLACK, (563, 429), 42, 20)
            if menu_shop:
                pygame.draw.rect(WG, gs.PINK_MUSIC, (988, 307, min(1, cash_ratio) * 145, 44))
                pygame.draw.circle(WG, gs.BLACK, (957, 325), 41, 20)

            # Update Slider
            volume_temp_color = gs.BLACK
            if 585 <= mouse_x <= 800 and 400 <= mouse_y <= 455:
                volume_temp = gs.calculate_volume(mouse_x)
                volume_temp_color = gs.WHITE
                if pygame.mouse.get_pressed()[0]:
                    volume = volume_temp
                    pygame.mixer.music.set_volume(volume / 100)
                    volume_temp_color = gs.BLACK

            # Draw Big Buttons
            if button_02_back.draw(surface=WG):
                menu_start = True
                menu_settings = False
            elif button_02_sources.draw(surface=WG):
                menu_licenses = True
                menu_settings = False
            elif button_02_shop.draw(surface=WG):
                menu_shop = not menu_shop
            elif button_02_how2play.draw(surface=WG):
                menu_info = True
                menu_settings = False

            # draw Mini Buttons
            if button_02_restore_TP.draw(surface=WG):
                restore_top_list()
            elif button_02_remove_TP.draw(surface=WG):
                gs.rename_file(old_name='score', new_name='score_backUp')
            elif button_02_controls.draw(surface=WG):
                menu_controls = True
                menu_settings = False

            #  Draw next / prev skin Buttons
            if button_02_p1_prev.draw(surface=WG):
                player1_skin_image_index = (player1_skin_image_index - 1) % len(img_list_player)
                player.update_image(player1_skin_image_index)
            elif button_02_p1_next.draw(surface=WG):
                player1_skin_image_index = (player1_skin_image_index + 1) % len(img_list_player)
                player.update_image(player1_skin_image_index)
            elif button_02_p2_prev.draw(surface=WG) and not menu_shop:
                player2_skin_image_index = (player2_skin_image_index - 1) % len(img_list_player)
                player2.update_image(player2_skin_image_index)
            elif button_02_p2_next.draw(surface=WG) and not menu_shop:
                player2_skin_image_index = (player2_skin_image_index + 1) % len(img_list_player)
                player2.update_image(player2_skin_image_index)
            elif button_02_random.draw(surface=WG):
                local_dl = len(img_list_player)
                if not menu_shop:
                    player1_skin_image_index = ((player1_skin_image_index + random.randint(1, 100)) % local_dl)
                    player2_skin_image_index = ((player2_skin_image_index + random.randint(1, 100)) % local_dl)
                else:
                    player1_skin_image_index = ((player1_skin_image_index + random.randint(1, 100)) % local_dl)

            # Change info text:
            if button_02_restore_TP.mouse_in():
                settings_text = "Restore Top List"
            elif button_02_remove_TP.mouse_in():
                settings_text = "Remove Top List"
            elif button_02_controls.mouse_in():
                settings_text = "Change player controls"
            elif button_02_music.mouse_in():
                settings_text = "Background music volume"
            elif button_02_sources.mouse_in():
                settings_text = "View sources"
            elif button_02_shop.mouse_in():
                settings_text = "Close SKINS shop" if menu_shop else "Open SKINS shop"
            elif button_02_back.mouse_in():
                settings_text = "Back to Main Menu"
            elif button_02_how2play.mouse_in():
                settings_text = "How To Play?"
            elif button_02_p1_prev.mouse_in():
                settings_text = "Previous skin"
            elif button_02_p1_next.mouse_in():
                settings_text = "Next skin"
            elif button_02_p2_prev.mouse_in() and not menu_shop:
                settings_text = "Previous skin"
            elif button_02_p2_next.mouse_in() and not menu_shop:
                settings_text = "Next skin"
            elif button_02_random.mouse_in() and not menu_shop:
                settings_text = "Random skin"
            elif button_02_buy_skin.mouse_in() and menu_shop:
                if not costumes[f"costume_{player1_skin_image_index}"]["purchased"]:
                    settings_text = "Buy champion SKIN"
                else:
                    settings_text = "You already own this SKIN"
            elif button_02_bank.mouse_in() and menu_shop:
                settings_text = "Your CA$H balance"
            else:
                settings_text = ""

            # Buy skin button - must be drawn before screen
            if menu_shop:
                if button_02_buy_skin.draw(surface=WG):
                    if not costumes[f"costume_{player1_skin_image_index}"]["purchased"] and menu_shop:
                        skin_price = int(costumes[f"costume_{player1_skin_image_index}"]["price"])
                        if money >= skin_price:
                            money -= skin_price
                            costumes[f"costume_{player1_skin_image_index}"]["purchased"] = True
                            sound_buy_skin.play()

            # Draw Screen
            WG.blit(image_menu_settings, rect_top_left)
            extra_image = image_menu_shop if menu_shop else image_menu_players
            WG.blit(extra_image, rect_top_right)

            # Open shop submenu
            if menu_shop:
                shop_text_price = "$ " + str(costumes[f"costume_{player1_skin_image_index}"]["price"])

                if not costumes[f"costume_{player1_skin_image_index}"]["purchased"]:
                    if money > 0:
                        cash_ratio = int(costumes[f"costume_{player1_skin_image_index}"]["price"]) / int(money)
                    else:
                        cash_ratio = 1
                    if money >= int(costumes[f"costume_{player1_skin_image_index}"]["price"]):
                        shop_text_color = gs.GREEN
                    else:
                        shop_text_color = gs.RED
                else:
                    shop_text_color = gs.GREEN
                    cash_ratio = 0

                gs.draw_text(surface=WG, text=f"$ {money}", x=1065, y=328, font=gs.FONT_36, color=gs.GREEN)
                gs.draw_text(surface=WG, text=shop_text_price, x=1065, y=412, font=gs.FONT_36, color=shop_text_color)

            # Draw player 1 skin assets
            player1_champ_name = str(costumes[f"costume_{player1_skin_image_index}"]["description"])
            gs.draw_text(surface=WG, text=player1_champ_name, x=1020, y=115, font=gs.FONT_28, color=gs.WHITE)
            gs.draw_image(surface=WG, image=img_list_player[player1_skin_image_index], x=1025, y=180, center=True)

            # Draw elements only for non-shop menu
            if not menu_shop:
                # draw player 1 asset
                gs.draw_text(surface=WG, text="P1:", font=gs.FONT_36, color=gs.YELLOW, y=115,
                             x=985 - gs.text_width(player1_champ_name, gs.FONT_28) // 2)

                # draw players 2 assets
                gs.draw_image(surface=WG, image=img_list_player[player2_skin_image_index], x=1025, y=370, center=True)
                player2_champ_name = str(costumes[f"costume_{player2_skin_image_index}"]["description"])
                gs.draw_text(surface=WG, text=player2_champ_name, x=1020, y=290, font=gs.FONT_28, color=gs.WHITE)
                gs.draw_text(surface=WG, text="P2:", font=gs.FONT_36, color=gs.BLUE, y=290,
                             x=985 - gs.text_width(player2_champ_name, gs.FONT_28) // 2)

                # draw lock images when skin not owned
                if not costumes[f"costume_{player1_skin_image_index}"]["purchased"]:
                    gs.draw_image(surface=WG, image=image_player_locked, x=1030, y=190, center=True)
                    skin_not_owned = True
                    purchase_text = "You don't own this skin"
                if not costumes[f"costume_{player2_skin_image_index}"]["purchased"]:
                    gs.draw_image(surface=WG, image=image_player_locked, x=1030, y=380, center=True)
                    skin_not_owned = True
                    purchase_text = "You don't own this skin"

                # draw skin info text:
                gs.draw_text(surface=WG, text=purchase_text, x=1025, y=445, font=gs.FONT_25, color=gs.WHITE)

            # Draw Text
            gs.draw_text(surface=WG, text=settings_text, x=930, y=530, font=gs.FONT_36, color=gs.WHITE)
            gs.draw_text(surface=WG, text=f'{volume}%', x=830, y=432, font=gs.FONT_36, color=gs.WHITE)
            gs.draw_text(surface=WG, text=f'({volume_temp}%)', x=830, y=460, font=gs.FONT_25, color=volume_temp_color)
            if settings_text:
                gs.draw_image(surface=WG, image=image_info, x=890 - gs.text_width(settings_text, gs.FONT_36) // 2,
                              y=530, center=True)

        # Open Controls Menu
        if menu_controls:
            WG.fill(color=gs.BLACK)

            # Save / Load / Clear stats
            controls_menu_text = ["SAVE", "LOAD", "CLEAR"]
            # controls_menu_typ  0: save,  1: load,  2: clear

            # Draw change buttons
            if button_03_back.draw(WG):
                menu_controls = False
                menu_settings = True
            if button_03_disk_prev.draw(WG):
                controls_menu_typ = (controls_menu_typ - 1) % 3
            if button_03_disk_next.draw(WG):
                controls_menu_typ = (controls_menu_typ + 1) % 3

            # Draw save data buttons
            if button_03_yellow.draw(WG):  # yellow button
                if controls_menu_typ == 0:
                    save_data_to_file(data=controls, slot_number=1)
                elif controls_menu_typ == 1:
                    loaded_controls = load_data_from_file(slot_number=1)
                    if loaded_controls:
                        controls = loaded_controls
                elif controls_menu_typ == 2:
                    controls = restart_data_file(slot_number=1)

            if button_03_pink.draw(WG):  # pink button
                if controls_menu_typ == 0:
                    save_data_to_file(data=controls, slot_number=2)
                elif controls_menu_typ == 1:
                    loaded_controls = load_data_from_file(slot_number=2)
                    if loaded_controls:
                        controls = loaded_controls
                elif controls_menu_typ == 2:
                    controls = restart_data_file(slot_number=2)

            if button_03_blue.draw(WG):  # blue button
                if controls_menu_typ == 0:
                    save_data_to_file(data=controls, slot_number=3)
                elif controls_menu_typ == 1:
                    loaded_controls = load_data_from_file(slot_number=3)
                    if loaded_controls:
                        controls = loaded_controls
                elif controls_menu_typ == 2:
                    controls = restart_data_file(slot_number=3)

            # Draw Screen Image
            WG.blit(image_menu_controls, rect_top_left)

            # Draw text
            gs.draw_text(surface=WG, text=controls_menu_text[controls_menu_typ], x=890, y=705, font=gs.FONT_36,
                         color=gs.WHITE, center=True)

            x_position = 800
            y_position = 133
            row_height = 87
            col_width = 225
            items_per_row = 5

            for index, (action, key) in enumerate(controls.items()):
                font = gs.create_font(size=55) if index < 10 else gs.create_font(size=40)
                color = gs.YELLOW_2 if index < 5 else gs.BLUE if index < 10 else gs.PINK
                if editing_key == action:
                    text = font.render("...", True, gs.RED)
                else:
                    text = font.render(f"{pygame.key.name(key)}".upper(), True, color)

                x = x_position + (index // items_per_row) * col_width
                y = y_position + (index % items_per_row) * row_height
                if index < 10:
                    WG.blit(text, (x - text.get_width() // 2, y))
                elif index == 10:
                    WG.blit(text, (160 - text.get_width() // 2, 385))
                elif index == 11:
                    WG.blit(text, (190 - text.get_width() // 2, 310))

        # Open info menu
        if menu_licenses:
            WG.fill(color=gs.BLACK)

            # Draw Buttons
            if button_06_00.draw(surface=WG):
                menu_settings = True
                menu_licenses = False
            elif button_06_01.draw(surface=WG):
                webbrowser.open(sources_link[licenses_index])
            elif button_06_02.draw(surface=WG):
                try:
                    subprocess.Popen(['start', 'excel', license_file_path], shell=True)
                    print("Excel successfully opened ")
                except Exception as e:
                    print(f"Error while opening Excel file: {e}")
            elif button_06_prev.draw(surface=WG):
                licenses_index = (licenses_index - 1) % licenses_max_index
            elif button_06_next.draw(surface=WG):
                licenses_index = (licenses_index + 1) % licenses_max_index

            # Draw Screen
            WG.blit(image_menu_licenses, rect_top_left)

            # Draw button description
            sources_text = ""
            buttons_sources = [button_06_00, button_06_01, button_06_02, button_06_prev, button_06_next]
            for button in buttons_sources:
                if button.mouse_in():
                    sources_text = button.description

            gs.draw_text(surface=WG, text=sources_text, x=400, y=635, font=gs.FONT_36)
            if sources_text:
                gs.draw_image(surface=WG, image=image_info, x=340 - gs.text_width(sources_text, gs.FONT_36) // 2, y=610)

            # Draw text
            gs.draw_text(surface=WG, text=f"{sources_typ[licenses_index]}", x=190, y=45, font=gs.FONT_50, center=False)
            link_length = len(sources_link[licenses_index])
            if link_length <= 40:
                gs.draw_text(surface=WG, text=f"{sources_link[licenses_index]}", x=190, y=160, font=gs.FONT_50,
                             center=False)
            else:
                gs.draw_text(surface=WG, text=f"{sources_link[licenses_index][:link_length // 2]}", x=190, y=150,
                             font=gs.FONT_32, center=False)
                gs.draw_text(surface=WG, text=f"{sources_link[licenses_index][link_length // 2:]}", x=190, y=185,
                             font=gs.FONT_32, center=False)
            gs.draw_text(surface=WG, text=f"{sources_name[licenses_index]}", x=220, y=280, font=gs.FONT_50,
                         center=False)

        if menu_info:
            WG.fill(color=gs.BLACK)

            # Draw Buttons - basic
            if button_04_prev.draw(surface=WG):
                info_image_index = (info_image_index - 1) % (len(img_list_how_to_play) - 1)
            elif button_04_next.draw(surface=WG):
                info_image_index = (info_image_index + 1) % (len(img_list_how_to_play) - 1)
            elif button_04_back.draw(surface=WG):
                menu_info = False
                menu_start = True

            # Draw images
            gs.draw_image(surface=WG, image=image_buttons_info_menu, x=755, y=165)
            gs.draw_image(surface=WG, image=img_list_how_to_play[-1], x=gs.WINDOW_WIDTH - 450, y=gs.WINDOW_HEIGHT - 450)
            gs.draw_image(surface=WG, image=img_list_how_to_play[info_image_index], x=90, y=160)

            # Draw text
            gs.draw_text(surface=WG, text="Feed The Mouse", font=gs.FONT_110, color=gs.YELLOW, x=550, y=110)
            gs.draw_text(surface=WG, text="3", font=gs.FONT_120, color=gs.WHITE, x=1055, y=115)

        # Draw cursor
        WG.blit(current_cursor, (mouse_x, mouse_y))

    if game_screen:
        # Check for game over
        if number_of_players == 1:
            if player.dead and not game_over:
                sound_gameOver.play()
                game_over = True
                if player.score > 0 and save_score:
                    ws.append([player.score, player.game_time,
                               player.cheese_amount + player.cheese_bonus_amount])
                    wb.save(excel_file_path)
                    save_score = False
        elif number_of_players == 2:
            if player.dead and player2.dead and not game_over:
                sound_gameOver.play()
                game_over = True
                if player.score > 0 and save_score:
                    ws.append([player.score, player.game_time,
                               player.cheese_amount + player.cheese_bonus_amount])
                    wb.save(excel_file_path)
                    save_score = False
                if player2.score > 0 and save_score2:
                    ws.append([player2.score, player2.game_time,
                               player2.cheese_amount + player2.cheese_bonus_amount])
                    wb.save(excel_file_path)
                    save_score2 = False

        # GameOver Screen
        if game_over:
            WG.fill(color=gs.BLACK)
            # Draw Player Stats
            if number_of_players == 1:
                # Draw player stats
                gs.draw_image(WG, image_game_over_one_player, 550, 70)
                gs.draw_image(WG, img_list_game_over[player.game_over], 725, 130, 0.625)
                gs.draw_text(WG, f'SCORE: {player.score:03d}', 750, 60, gs.FONT_75, center=False)
                gs.draw_text(WG, f'{player.cheese_amount:02d}', 625, 70, center=False)
                gs.draw_text(WG, f'{player.cheese_bonus_amount:02d}', 625, 150, center=False)
                gs.draw_text(WG, gs.format_ratio(player.cheese_fake_amount, game_state.TOTAL_FAKE, True), 620,
                             235, center=False)
                gs.draw_text(WG, gs.format_time(player.game_time), 620, 320, center=False)
                gs.draw_text(WG, gs.format_ratio(player.cheese_amount + player.cheese_bonus_amount,
                                                 game_state.TOTAL_CHEESE), 620, 420, center=False)
                # Draw text
                gs.draw_text_double(WG, 'GAME', 90, 60, 0, 7, 5,
                                    gs.FONT3_165, gs.RED, gs.FONT3_175, gs.WHITE, [0, 90, 110, 120])
                gs.draw_text_double(WG, 'OVER', 90, 260, 0, 7, 5,
                                    gs.FONT3_165, gs.RED, gs.FONT3_175, gs.WHITE, [0, 100, 100, 75])
                gs.draw_text(WG, 'Press ENTER to play again', 75, 470, gs.FONT_36, center=False)

            elif number_of_players == 2:
                main_img_index = 5
                main_text = "      DRAW       "
                if player.score > player2.score:
                    main_img_index = 4
                    main_text = "Player 1 WIN!"
                elif player.score < player2.score:
                    main_img_index = 3
                    main_text = "Player 2 WIN!"

                gs.draw_image(WG, img_list_game_over[main_img_index], 300, 125, 0.75)

                # Draw player1 stats
                extra_x, extra_y = 25, 70
                gs.draw_text(WG, "PLAYER 1", 50 + extra_x, -10 + extra_y, gs.FONT_50, center=False)
                gs.draw_image(WG, image_game_over_two_players, 30 + extra_x, 75 + extra_y)
                gs.draw_text(WG, f'{player.score:<4}', 120 + extra_x, 80 + extra_y, center=False)
                gs.draw_text(WG, f'{player.cheese_amount:02d}', 120 + extra_x, 145 + extra_y, center=False)
                gs.draw_text(WG, f'{player.cheese_bonus_amount:02d}', 120 + extra_x, 215 + extra_y, center=False)
                gs.draw_text(WG, gs.format_ratio(player.cheese_fake_amount, game_state.TOTAL_FAKE // 2, True),
                             120 + extra_x, 290 + extra_y, center=False)
                gs.draw_text(WG, gs.format_time(player.game_time), 120 + extra_x, 370 + extra_y, center=False)
                gs.draw_text(WG, gs.format_ratio(player.cheese_amount + player.cheese_bonus_amount,
                                                 game_state.TOTAL_CHEESE // 2), 120 + extra_x, 455 + extra_y,
                             center=False)

                # Draw player2 stats
                extra_x, extra_y = 950, 70
                gs.draw_text(WG, "PLAYER 2", extra_x, -10 + extra_y, gs.FONT_50, center=False)
                gs.draw_image(WG, image_game_over_two_players, 30 + extra_x, 75 + extra_y, center=False)
                gs.draw_text(WG, f'{player2.score:<4}', 120 + extra_x, 80 + extra_y, center=False)
                gs.draw_text(WG, f'{player2.cheese_amount:02d}', 120 + extra_x, 145 + extra_y, center=False)
                gs.draw_text(WG, f'{player2.cheese_bonus_amount:02d}', 120 + extra_x, 215 + extra_y, center=False)
                gs.draw_text(WG, gs.format_ratio(player2.cheese_fake_amount, game_state.TOTAL_FAKE // 2, True),
                             120 + extra_x, 290 + extra_y, center=False)
                gs.draw_text(WG, gs.format_time(player2.game_time), 120 + extra_x, 370 + extra_y, center=False)
                gs.draw_text(WG, gs.format_ratio(player2.cheese_amount + player2.cheese_bonus_amount,
                                                 game_state.TOTAL_CHEESE // 2), 120 + extra_x, 455 + extra_y,
                             center=False)

                # Drw Text
                gs.draw_text(WG, main_text, 385, 50, gs.FONT_75, gs.YELLOW, center=False)

            # Draw Buttons
            if button_05_00.draw(surface=WG):  # Button - Restart Game
                menu_screen = True
                menu_start = True
                game_over = False
                game_screen = False
                game_paused = False
                restart_game()

            if button_05_01.draw(surface=WG):  # Button - Back to Main Menu
                game_over = False
                game_paused = False
                restart_game()

            if button_05_02.draw(surface=WG):  # Button - Quit Game
                game_running = False
                # Add Money
                money += player.score // 15 + 2 * player.cheese_amount
                money += player2.score // 15 + 2 * player2.cheese_amount

            # Draw Cursor
            WG.blit(current_cursor, (mouse_x, mouse_y))

            # Restart  save_score FLAGS
            save_score = True
            save_score2 = True

        # Game
        else:
            # Draw assets (player, cheese, items)
            WG.fill(color=gs.BLACK)
            cheese_group.draw(WG)
            item_group.draw(WG)
            trap_group.draw(WG)
            player.draw(WG)
            if debug_mouse_pos:
                WG.blit(current_cursor, (mouse_x, mouse_y))

            # Collision checker
            if debug_collision:
                pygame.draw.rect(WG, gs.WHITE, player.rect, 1)
                pygame.draw.rect(WG, gs.PINK, player.rect_image, 2)
                for item in item_group:
                    pygame.draw.rect(WG, gs.RED, item.rect, 3)
                    pygame.draw.rect(WG, gs.GREEN, item.rect_col, 2)
                for cheese in cheese_group:
                    pygame.draw.rect(WG, gs.YELLOW, cheese.rect, 2)
                for trap in trap_group:
                    pygame.draw.rect(WG, gs.RED, trap.rect, 2)
                    for rect in trap.collision_rects:
                        if not trap.locked:
                            pygame.draw.rect(WG, gs.YELLOW_2, rect, 2)

            if number_of_players == 2:
                player.draw_hud(surface=WG, offset_x=0, bar_width=240, one_player=False)
                player2.draw(WG)
                player2.draw_hud(surface=WG, offset_x=920, bar_width=240, one_player=False)
            else:
                player.draw_hud(surface=WG, offset_x=0, bar_width=240, one_player=True)

            # Show cheese speed
            if show_cheese_speed and number_of_players == 1:
                gs.draw_text(surface=WG, text=f'Cheese speed: {cheese_speed}', x=1080, y=670,
                             font=gs.FONT_25, color=gs.ORANGE)
                gs.draw_text(surface=WG, text=f'Poison: {player.poison}%', x=1115, y=700,
                             font=gs.FONT_25, color=gs.GREEN)
                gs.draw_text(surface=WG, text=f'Miss in row: {player.game_state.MISS_IN_ROW}', x=1090, y=730,
                             font=gs.FONT_25, color=gs.RED)

            # Pause Game
            if game_paused:
                pygame.draw.rect(WG, gs.BLACK, (0, 0, gs.WINDOW_WIDTH, gs.HUD2))
                if pause_game_image_index > 1:
                    gs.draw_image(WG, img_list_pause_game[pause_game_image_index], gs.WINDOW_WIDTH - 400, 225,
                                  0.50)
                    gs.draw_text_double(WG, 'PAUSED', 350, 155, 75, 7, 5,
                                        gs.FONT2_165, gs.RED, gs.FONT2_175, gs.WHITE)
                else:
                    gs.draw_image(WG, img_list_pause_game[pause_game_image_index], 250, 0, 0.95)

            # Update game
            if not game_paused:
                cheese_group.update()
                player.move(controls)
                player.update(cheese_group)
                item_group.update(player)
                trap_group.update(player)

                # update cheese speed
                if time.time() - last_update_time >= gs.CHEESE_SPEED_cooldown:
                    for cheese in cheese_group:
                        cheese.speed += gs.CHEESE_SPEED_scale

                    cheese_speed = round(cheese_group.sprites()[0].speed, 2)
                    last_update_time = time.time()

                # update items speed
                if time.time() - last_update_time >= gs.CHEESE_SPEED_cooldown:
                    for item in item_group:
                        item.speed += gs.CHEESE_SPEED_scale

                if number_of_players == 2:
                    player2.move(controls)
                    player2.update(cheese_group)
                    item_group.update(player2)
                    trap_group.update(player2)
                    player.update_time(player2)
                    player2.update_time(player)

    # Update display and tick clock
    pygame.display.update()
    clock.tick(FPS)
# endregion


# region 8. End the game
game_data = {"money": money, "costumes": costumes}
save_data_to_file(data=game_data, slot_number=1, filename_prefix='./config/data/costume_shop_')

settings_data = {"player_1": player1_skin_image_index, "player_2": player2_skin_image_index, "volume": volume}
save_data_to_file(data=settings_data, slot_number=0, filename_prefix='./config/data/settings_')

pygame.quit()
sys.exit()
# endregion
